#!/usr/bin/env python3
"""
Recompute Clarity + diary merge and compare to the enriched workbook on disk.

Exits 0 if At Sea / At Work / Glucose Issue match row-by-row; otherwise prints
mismatches and exits 1.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

import enrich_clarity as ec  # noqa: E402
from diary_load import load_clean_diary  # noqa: E402
from input_resolve import find_unique_by_prefix, find_unique_clarity_raw  # noqa: E402
from term import Term, print_run_footer, print_run_header, use_color  # noqa: E402

DIARY_PREFIX = "SEAGM"


def _norm_cell(x) -> str:
    if x is None:
        return ""
    if isinstance(x, float) and pd.isna(x):
        return ""
    if isinstance(x, (float, np.floating)) and not pd.isna(x):
        if float(x).is_integer():
            return str(int(x))
    s = str(x).strip()
    if s.lower() == "nan":
        return ""
    return s


def _norm_diary_value(col: str, x) -> str:
    """Match enrich output semantics: bools and 0/1 as TRUE/FALSE; strings uppercased for sea/work."""
    if col in ("At Sea", "At Work"):
        if isinstance(x, (bool, np.bool_)):
            return "TRUE" if x else "FALSE"
        if isinstance(x, (int, np.integer)):
            if x == 1:
                return "TRUE"
            if x == 0:
                return "FALSE"
        if isinstance(x, (float, np.floating)) and not pd.isna(x):
            if x == 1.0:
                return "TRUE"
            if x == 0.0:
                return "FALSE"
        u = _norm_cell(x).upper()
        if u in ("TRUE", "FALSE"):
            return u
        return _norm_cell(x)
    return _norm_cell(x)


@dataclass
class VerifyOutcome:
    exit_code: int
    total_rows: int = 0
    ok_rows: int = 0
    bad_rows: int = 0
    row_count_mismatch: bool = False
    mismatch_lines: list[str] = field(default_factory=list)


def _sheet_compare_issues(
    *,
    sheet_name: str,
    expected: pd.DataFrame,
    actual: pd.DataFrame,
    max_lines: int,
    progress_cb=None,
) -> tuple[int, list[str]]:
    issues = 0
    lines: list[str] = []
    if len(expected) != len(actual):
        issues += 1
        lines.append(
            f"  [{sheet_name}] row count mismatch: expected {len(expected)}, got {len(actual)}"
        )

    if list(expected.columns) != list(actual.columns):
        issues += 1
        lines.append(
            f"  [{sheet_name}] column mismatch: expected {list(expected.columns)!r}, got {list(actual.columns)!r}"
        )

    common_cols = [c for c in expected.columns if c in actual.columns]
    nrows = min(len(expected), len(actual))
    for i in range(nrows):
        if progress_cb is not None:
            progress_cb(i + 1, nrows)
        for col in common_cols:
            if col in ec.DIARY_COLS:
                e = _norm_diary_value(col, expected.iloc[i][col])
                a = _norm_diary_value(col, actual.iloc[i][col])
            else:
                e = _norm_cell(expected.iloc[i][col])
                a = _norm_cell(actual.iloc[i][col])
            if e != a:
                issues += 1
                if len(lines) < max_lines:
                    lines.append(
                        f"  [{sheet_name}] row {i} {col!r}: expected {e!r}, file has {a!r}"
                    )
    return issues, lines


def _summary_issues(
    *,
    expected: pd.DataFrame,
    actual: pd.DataFrame,
    max_lines: int,
    progress_cb=None,
) -> tuple[int, list[str]]:
    issues = 0
    lines: list[str] = []
    rows, cols = expected.shape
    for r in range(rows):
        if progress_cb is not None:
            progress_cb(r + 1, rows)
        for c in range(cols):
            e = _norm_cell(expected.iat[r, c])
            if e == "":
                continue
            a = ""
            if r < actual.shape[0] and c < actual.shape[1]:
                a = _norm_cell(actual.iat[r, c])
            if e != a:
                issues += 1
                if len(lines) < max_lines:
                    lines.append(
                        f"  [{ec.SHEET_SUMMARY}] cell R{r + 1}C{c + 1}: expected {e!r}, file has {a!r}"
                    )
    return issues, lines


def _read_sheet_raw_values(
    *,
    workbook_path: Path,
    sheet_name: str,
    rows: int,
    cols: int,
    progress_cb=None,
) -> pd.DataFrame:
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb[sheet_name]
        raw: list[list[object]] = []
        for r in range(1, rows + 1):
            if progress_cb is not None:
                progress_cb(r, rows)
            row: list[object] = []
            for c in range(1, cols + 1):
                row.append(ws.cell(row=r, column=c).value)
            raw.append(row)
        return pd.DataFrame(raw)
    finally:
        wb.close()


def verify(
    *,
    diary_path: Path,
    clarity_raw_path: Path,
    enriched_path: Path,
    max_report: int,
    term: Term,
    show_progress: bool,
) -> VerifyOutcome:
    def say(msg: str = "") -> None:
        print(msg)

    print_run_header(term, "Clarity × diary enrich verify")
    say(f"  {term.dim('Diary:')}     {diary_path.name}")
    say(f"  {term.dim('Clarity:')}   {clarity_raw_path.name}")
    say(f"  {term.dim('Enriched:')}  {enriched_path.name}")
    say()

    say(term.dim("  Recomputing merge from diary + raw Clarity…"))
    diary = load_clean_diary(str(diary_path))
    clarity = pd.read_excel(clarity_raw_path, engine="openpyxl")
    for c in ec.DIARY_COLS:
        if c not in clarity.columns:
            clarity[c] = ""

    expected = ec.enrich_clarity(clarity, diary)
    expected_filtered = ec.build_filtered_sheets(expected)
    expected_summary = ec.build_summary_sheet(
        total=len(expected),
        on_land=len(expected_filtered[ec.SHEET_ON_LAND]),
        at_sea=len(expected_filtered[ec.SHEET_AT_SEA]),
        at_sea_at_work=len(expected_filtered[ec.SHEET_AT_SEA_AT_WORK]),
    )
    say(term.dim("  Loading enriched workbook…"))
    workbook = pd.ExcelFile(enriched_path, engine="openpyxl")
    if ec.SHEET_ENRICHED in workbook.sheet_names:
        actual = pd.read_excel(workbook, sheet_name=ec.SHEET_ENRICHED)
    else:
        actual = pd.read_excel(workbook, sheet_name=workbook.sheet_names[0])

    if len(expected) != len(actual):
        say()
        say(term.red(term.bold("  ✗ FAILED")))
        say(
            term.red(
                f"  Row count mismatch: expected {len(expected)} (raw Clarity), "
                f"got {len(actual)} in enriched file."
            )
        )
        print_run_footer(term)
        return VerifyOutcome(exit_code=1, row_count_mismatch=True)

    total = len(expected)
    bad_rows = 0
    lines_printed = 0
    mismatch_lines: list[str] = []
    last_pct = -1
    progress_printed = False

    for i in range(total):
        row_issues: list[tuple[str, str, str]] = []
        for col in ec.DIARY_COLS:
            e = _norm_diary_value(col, expected.iloc[i][col])
            a = _norm_diary_value(col, actual.iloc[i][col])
            if e != a:
                row_issues.append((col, e, a))
        if row_issues:
            bad_rows += 1
            ts = expected.iloc[i].get(ec.CLARITY_TS_COL, "")
            for col, e, a in row_issues:
                if lines_printed < max_report:
                    mismatch_lines.append(
                        f"  row {i} ts={ts!r} {col!r}: expected {e!r}, file has {a!r}"
                    )
                    lines_printed += 1

        # Progress only on a real TTY so logs (make, pipes) stay short.
        # Redraw in-place: CR + clear line (ANSI) avoids stacking lines in normal terminals.
        if show_progress and total >= 50 and sys.stdout.isatty():
            pct = 100 * (i + 1) // total
            if pct != last_pct:
                bar_w = 24
                filled = int(bar_w * (i + 1) / total)
                bar = "█" * filled + "░" * (bar_w - filled)
                if term.enabled:
                    msg = f"  {term.cyan('Comparing')}{term.dim('…')} [{bar}] {pct:>3}%"
                else:
                    msg = f"  Comparing… [{bar}] {pct:>3}%"
                sys.stdout.write("\r\033[2K" + msg)
                sys.stdout.flush()
                last_pct = pct
                progress_printed = True

    if progress_printed:
        print()

    ok_rows = total - bad_rows
    pct_ok = (100.0 * ok_rows / total) if total else 100.0

    say(term.dim("  Validating additional sheets…"))
    inline_detail = show_progress and sys.stdout.isatty()

    def draw_detail(msg: str) -> None:
        if inline_detail:
            sys.stdout.write("\r\033[2K" + msg)
            sys.stdout.flush()
        else:
            say(msg)

    def finish_detail_line() -> None:
        if inline_detail:
            print()

    def format_detail_progress(label: str, pct: int) -> str:
        bar_w = 24
        filled = int(bar_w * pct / 100)
        bar = "█" * filled + "░" * (bar_w - filled)
        if term.enabled:
            return f"      {term.cyan(label)}{term.dim('…')} [{bar}] {pct:>3}%"
        return f"      {label}... [{bar}] {pct:>3}%"

    extra_issues = 0
    for sheet_name in (ec.SHEET_AT_SEA, ec.SHEET_AT_SEA_AT_WORK, ec.SHEET_ON_LAND):
        say(term.dim(f"    - {sheet_name}"))
        if sheet_name not in workbook.sheet_names:
            extra_issues += 1
            if lines_printed < max_report:
                mismatch_lines.append(f"  Missing required sheet: {sheet_name!r}")
                lines_printed += 1
            continue
        draw_detail(term.dim("      loading rows…"))
        act = pd.read_excel(workbook, sheet_name=sheet_name)
        sheet_cmp_last = -1

        def sheet_compare_progress(done: int, total_rows: int) -> None:
            nonlocal sheet_cmp_last
            if total_rows <= 0:
                return
            pct = int((100 * done) / total_rows)
            if pct != sheet_cmp_last:
                draw_detail(format_detail_progress("Comparing rows", pct))
                sheet_cmp_last = pct

        rem = max(max_report - lines_printed, 0)
        issues, lines = _sheet_compare_issues(
            sheet_name=sheet_name,
            expected=expected_filtered[sheet_name],
            actual=act,
            max_lines=rem,
            progress_cb=sheet_compare_progress if inline_detail else None,
        )
        if inline_detail and sheet_cmp_last < 100:
            draw_detail(format_detail_progress("Comparing rows", 100))
        finish_detail_line()
        extra_issues += issues
        mismatch_lines.extend(lines)
        lines_printed += len(lines)

    if ec.SHEET_SUMMARY not in workbook.sheet_names:
        say(term.dim(f"    - {ec.SHEET_SUMMARY}"))
        extra_issues += 1
        if lines_printed < max_report:
            mismatch_lines.append(f"  Missing required sheet: {ec.SHEET_SUMMARY!r}")
            lines_printed += 1
    else:
        say(term.dim(f"    - {ec.SHEET_SUMMARY}"))
        summary_load_last = -1

        def summary_load_progress(done: int, total_rows: int) -> None:
            nonlocal summary_load_last
            if total_rows <= 0:
                return
            pct = int((100 * done) / total_rows)
            if pct != summary_load_last:
                draw_detail(format_detail_progress("Loading cells", pct))
                summary_load_last = pct

        act_summary = _read_sheet_raw_values(
            workbook_path=enriched_path,
            sheet_name=ec.SHEET_SUMMARY,
            rows=expected_summary.shape[0],
            cols=expected_summary.shape[1],
            progress_cb=summary_load_progress if inline_detail else None,
        )
        summary_cmp_last = -1

        def summary_compare_progress(done: int, total_rows: int) -> None:
            nonlocal summary_cmp_last
            if total_rows <= 0:
                return
            pct = int((100 * done) / total_rows)
            if pct != summary_cmp_last:
                draw_detail(format_detail_progress("Comparing cells", pct))
                summary_cmp_last = pct

        rem = max(max_report - lines_printed, 0)
        issues, lines = _summary_issues(
            expected=expected_summary,
            actual=act_summary,
            max_lines=rem,
            progress_cb=summary_compare_progress if inline_detail else None,
        )
        if inline_detail and summary_cmp_last < 100:
            draw_detail(format_detail_progress("Comparing cells", 100))
        finish_detail_line()
        extra_issues += issues
        mismatch_lines.extend(lines)
        lines_printed += len(lines)

    total_issues = bad_rows + extra_issues

    say()
    say(f"  {term.dim('Rows checked:')}  {total:,}")
    say(f"  {term.dim('Extra checks:')}  filtered sheets + summary")
    ok_part = term.green(f"{ok_rows:,} ok") if total_issues == 0 else f"{ok_rows:,} ok"
    fail_part = term.red(f"{total_issues:,} failed") if total_issues else term.dim("0 failed")
    say(f"  {term.dim('Match rate:')}   {pct_ok:>6.2f}%  ({ok_part}, {fail_part})")
    say()

    if total_issues:
        say(term.yellow(term.bold(f"  Sample mismatches (up to {max_report} lines):")))
        for ln in mismatch_lines:
            say(term.red(ln))
        if lines_printed >= max_report:
            say(term.dim(f"  … (stopped after {max_report} lines; more may exist)"))
        say()
        say(term.red(term.bold("  ✗ FAILED")))
        print_run_footer(term)
        return VerifyOutcome(
            exit_code=1,
            total_rows=total,
            ok_rows=ok_rows,
            bad_rows=bad_rows,
            mismatch_lines=mismatch_lines,
        )

    say(term.green(term.bold("  ✓ PASSED")))
    say(term.green(f"  Enriched file matches recomputed diary merge ({total:,} rows)."))
    print_run_footer(term)
    return VerifyOutcome(exit_code=0, total_rows=total, ok_rows=ok_rows, bad_rows=0)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input-dir", default="input", help="Diary + raw Clarity (default: input)")
    p.add_argument("--output-dir", default="output", help="Enriched workbook (default: output)")
    p.add_argument("--diary", "-d", default=None, help="Diary .xlsx (default: discover SEAGM*)")
    p.add_argument("--clarity", "-c", default=None, help="Raw Clarity .xlsx (default: Clarity_Export*, not *_diary)")
    p.add_argument("--enriched", "-e", default=None, help="Enriched .xlsx (default: <clarity_stem>_diary.xlsx in output-dir)")
    p.add_argument(
        "--max-report",
        type=int,
        default=30,
        metavar="N",
        help="Max mismatch lines to print (default: 30)",
    )
    p.add_argument("--no-color", action="store_true", help="Disable ANSI colors")
    p.add_argument(
        "--no-progress",
        action="store_true",
        help="Disable the comparing progress bar",
    )
    args = p.parse_args()

    term = Term(use_color(no_color_flag=args.no_color))

    diary_path = (
        Path(args.diary)
        if args.diary is not None
        else find_unique_by_prefix(args.input_dir, DIARY_PREFIX, label="Diary")
    )
    clarity_raw_path = (
        Path(args.clarity)
        if args.clarity is not None
        else find_unique_clarity_raw(args.input_dir)
    )

    if args.enriched is not None:
        enriched_path = Path(args.enriched)
    else:
        enriched_path = Path(args.output_dir) / f"{clarity_raw_path.stem}_diary{clarity_raw_path.suffix}"

    if not enriched_path.is_file():
        print(term.red(f"Enriched file not found: {enriched_path}"), file=sys.stderr)
        return 2

    outcome = verify(
        diary_path=diary_path,
        clarity_raw_path=clarity_raw_path,
        enriched_path=enriched_path,
        max_report=args.max_report,
        term=term,
        show_progress=not args.no_progress,
    )
    return outcome.exit_code


if __name__ == "__main__":
    raise SystemExit(main())
